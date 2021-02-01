import requests
from bs4 import BeautifulSoup
import time
from urllib.request import Request
from urllib.request import urlopen
import urllib
import sys
import openpyxl
import os
import openpyxl
from datetime import datetime

import bs4, webbrowser, time, requests, pprint, getpass
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from datetime import datetime
from selenium.webdriver.firefox.options import Options
from selenium.common.exceptions import NoSuchElementException

def scroll_and_click(navegador):
    for i in range(10): # adjust integer value for need
        # you can change right side number for scroll convenience or destination 
        navegador.execute_script("window.scrollBy(0, 10)")
        time.sleep(.5)
        try:
            navegador.find_element_by_css_selector('#root > div > div.gtm-page-content > div > div.PageContainer__TabContainer-sc-1ae9mgv-5.nlBCd > div > div.PageContainer__ProductsContainer-sc-1ae9mgv-1.bmNqHs > div.WidthLimiter__Root-sjrjtk-0.gIbRfg > div > div > button > div').click()
            print('----->Ahora si')
        except NoSuchElementException:           
            print('Aun no')
            continue
        # you can change time integer to float or remove    

    
    
now = datetime.now()
date_time = now.strftime("%Y-%d-%m %H-%M Hs")
    
link_categories = 'https://www.capterra.com/categories'    
diccionario = {}
diccionario[link_categories]={}    
user_agent = {'User-agent': 'Mozilla/5.0','Accept-Language' : 'en-US,en;q=0.8'}

#============================CATEGORIAS
# link='https://planning.lacity.org/pdiscaseinfo/search/encoded/MjIxMDAz0'
# link = 'https://planning.lacity.org/pdiscaseinfo/search/encoded/MjE5ODU30'
# link ='https://planning.lacity.org/pdiscaseinfo/search/encoded/MjIwNzA10'
archivo = open(r'J:\Emma\Upwork\Gobierno\links.txt','r')
links = archivo.readlines()


    
wb = openpyxl.Workbook() # Abre el excel y toma solamente datos, si hay formulas solo toma el valor.
wb.active=0
hoja = wb.active
fila = 2 
datos = {}
PDFs = {}   

for link in links:
    
    # path_driver =r'C:\chromedriver.exe'
    path_driver = r'C:\geckodriver.exe'
    opciones = Options()
    opciones.headless = True
    navegador = webdriver.Firefox(executable_path=path_driver,options=opciones) 
    
    datos[str(fila)]= {}
    PDFs[str(fila)] = {}
    navegador.get(link)
    navegador.maximize_window()
    time.sleep(1.5)  
    scroll_and_click(navegador)
    home = navegador.window_handles[0]
    
    paginas = []
    for i in range(1,10):
        try:
            paginas.append(navegador.find_element_by_xpath('/html/body/app-root/main/div/app-search/mat-tab-group/div/mat-tab-body[1]/div/div/div[2]/div[1]/div/dx-data-grid/div/div[11]/div/div[{}]'.format(i)))
        except NoSuchElementException:
            break
    


    if not paginas:
        for x in range(2,22):
            try:
                tab = navegador.find_element_by_xpath('/html/body/app-root/main/div/app-search/mat-tab-group/div/mat-tab-body[1]/div/div/div[2]/div[1]/div/dx-data-grid/div/div[6]/div/table/tbody/tr[{}]/td[5]/a'.format(x))
                descripcion = str(navegador.find_element_by_xpath('/html/body/app-root/main/div/app-search/mat-tab-group/div/mat-tab-body[1]/div/div/div[2]/div[1]/div/dx-data-grid/div/div[6]/div/table/tbody/tr[{}]/td[2]'.format(x)).text)
                descripcion = descripcion.replace(':','')
                fecha = str(navegador.find_element_by_xpath('/html/body/app-root/main/div/app-search/mat-tab-group/div/mat-tab-body[1]/div/div/div[2]/div[1]/div/dx-data-grid/div/div[6]/div/table/tbody/tr[{}]/td[3]'.format(x)).text)
                fecha = fecha.replace('/','-')
                descripcion = descripcion+'-'+fecha
                tab.click()
                navegador.switch_to.window(navegador.window_handles[-1])
                time.sleep(3)
                pdf_url = navegador.current_url
                time.sleep(3)
                navegador.close()
                navegador.switch_to.window(home)
                time.sleep(3)
                if descripcion in PDFs[str(fila)].keys():
                    descripcion = descripcion+'-v2'
                PDFs[str(fila)][descripcion] = pdf_url
                print(descripcion,'      ',pdf_url)
            except NoSuchElementException:
                print('-----pasando 1')
                continue
    else:
        for pagina in range(1,len(paginas)+ 1):
            navegador.find_element_by_xpath('/html/body/app-root/main/div/app-search/mat-tab-group/div/mat-tab-body[1]/div/div/div[2]/div[1]/div/dx-data-grid/div/div[11]/div/div[{}]'.format(pagina)).click()
            time.sleep(3)
            for x in range(2,22):
                try:
                    
                    tab = navegador.find_element_by_xpath('/html/body/app-root/main/div/app-search/mat-tab-group/div/mat-tab-body[1]/div/div/div[2]/div[1]/div/dx-data-grid/div/div[6]/div/table/tbody/tr[{}]/td[5]/a'.format(x))
                    descripcion = str(navegador.find_element_by_xpath('/html/body/app-root/main/div/app-search/mat-tab-group/div/mat-tab-body[1]/div/div/div[2]/div[1]/div/dx-data-grid/div/div[6]/div/table/tbody/tr[{}]/td[2]'.format(x)).text)
                    descripcion = descripcion.replace(':','')
                    fecha = str(navegador.find_element_by_xpath('/html/body/app-root/main/div/app-search/mat-tab-group/div/mat-tab-body[1]/div/div/div[2]/div[1]/div/dx-data-grid/div/div[6]/div/table/tbody/tr[{}]/td[3]'.format(x)).text)
                    fecha = fecha.replace('/','-')
                    descripcion = descripcion+'-'+fecha
                    tab.click()
                    navegador.switch_to.window(navegador.window_handles[-1])
                    time.sleep(3)
                    pdf_url = navegador.current_url
                    time.sleep(3)
                    navegador.close()
                    navegador.switch_to.window(home)
                    time.sleep(3)
                    if descripcion in PDFs[str(fila)].keys():
                        descripcion = descripcion+'-v2'
                    PDFs[str(fila)][descripcion] = pdf_url
                    print(descripcion,'      ',pdf_url)
                except NoSuchElementException as e :
                    print(e)
                    continue
    
    pagesource = navegador.page_source
    sopa = BeautifulSoup(pagesource,'html.parser')
    navegador.quit()

    negrita = sopa.find_all('div',{'class':'rowData'})
    for x in range(len(negrita)):
        try:
            if x == len(negrita)- 1:
               tag = str(negrita[20].text)
               tag,valor = tag.split('  ')[0],tag.split('  ')[1]
            else:
                tag = negrita[x].find('div',{'class':'title'}).text
                valor = negrita[x].find('div',{'class':'data'}).text
            
            if str(valor) == 'None':
                valor = ''
            else:
                valor = str(valor).replace('\t','')            
            datos[str(fila)][str(tag)] = valor
            
            if tag == 'Case on Hold?:':
                for indice in range(len(negrita)):
                    try:
                        primary_Adress = negrita[indice].find('div',{'class':'pdisPropertyInfo'})
                        if str(primary_Adress) =='None':
                            continue
                        else:
                            cuadro = primary_Adress.find('table')
                            encabezados = cuadro.find_all('th')
                            valores = cuadro.find_all('td')
                            
                            for x in range(3):
                                tag = encabezados[x].text
                                valor = valores[x].text
                                if str(valor) == 'None':
                                    valor = ''
                                else:
                                    valor = str(valor).replace('\t','')                          
                                datos[str(fila)][str(tag)] = valor
                                
                    except NoSuchElementException:
                        continue
        except AttributeError:
            continue
    
    
    link = link.replace('\n','')
    link = link.replace('\t','')
    id_ = link.split('/')
    id_ = id_[len(id_) - 1]
    carpeta = os.path.join(os.getcwd(),'PDFs_id_{}'.format(id_))
    
    if PDFs:
        os.mkdir(carpeta)
        for pdf,url in PDFs[str(fila)].items():
            with open(os.path.join(carpeta,'PDF_{}.pdf'.format(pdf)),'wb') as arch:
                if str(PDFs[str(fila)][pdf]) != 'about:blank':
                    file = requests.get(url)
                    arch.write(file.content)
                else:
                    texto = 'The link was down, no pdf available'
                    arch.write(texto.encode())
    
            
    for indice,titulo in enumerate(datos['2'].keys()):
        hoja.cell(row=1,column=indice + 1).value = titulo
    
    for columna,data in enumerate(datos[str(fila)].values()):
        hoja.cell(row=fila,column=columna + 1).value = data

    wb.save(os.path.join(os.getcwd(),'Output_{}.xlsx'.format(date_time)))
    fila+=1

wb.close()


        